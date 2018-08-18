from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# ws.title = "PyBay rocks!!"
# ws.sheet_properties.tabColor = "1072BA"
# source = wb.active
# target = wb.copy_worksheet(source)

# wb.save("pybayrox.xlsx")


# ws['A1'] = "Annie Easley"
# d = ws.cell(row=1, column=2, value="Katherine Johnson")
# print(d)
# print(d.value)

# from openpyxl.drawing.image import Image
# img = Image('katherinejohnson.jpeg')
# # add to worksheet and anchor next to cells
# ws.add_image(img, 'A1')

# ws["A1"] = "=SUM(8, 8)"
# wb.save("formula.xlsx")

# # Rows can also be appended
# ws.append(["Grace Hopper", "Megan Smith", "Ada Lovelace"]) #next row, A2, B2, C2

# Save the file
wb.save("sample.xlsx")



