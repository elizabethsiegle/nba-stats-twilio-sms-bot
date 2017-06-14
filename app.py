from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from openpyxl import load_workbook, Workbook 


app = Flask(__name__)
@app.route('/', methods=['GET', 'POST'])

def parseExcel():
	excelfile = 'nba_stats_two_sheets.xlsx'
	typomsg = "send first and last name(s), stat, or check for typos!"
	statmsg = "Type 1-2 players' names(first + last, players separated by a space), then a stat(GP,W,L,MIN,PTS,FG%,3P%,FT%,REB,AST,STL,BLK)"
	msg = request.form['Body'].lower() #check that it's lowercase
	list_of_players = []
	list_of_stats = []
	#make dictionary: key = stat, value=column of Excel spreadsheet (column B is avg minutes played)
	
	stat_dict = { 
		"age":"B", "gp":"C","w":"D","l":"E","min":"F","pts":"G",
		"fgm":"H","fga":"I","fg%":"J",
		"3pm":"K","3pa":"L","ftm":"M","fta":"N",
		"ft%":"O","oreb":"P","dreb":"Q","reb":"R","ast":"S","tov":"T", "stl": "U",
		"blk": "V", "pf": "W", "dd2": "X", "td3": "Y"
	}
	wb = load_workbook(excelfile)
	reg_season = True # set boolean, will change depending on user input
	if msg == "play":
		ret = MessagingResponse().message("type \'a\' for regular season or \'b\' for playoffs")
	elif msg == 'a':
		ret = MessagingResponse().message(statmsg)
	elif msg == 'b':
		reg_season = False
		ret = MessagingResponse().message(statmsg)
	else:
		if reg_season: # playoffs or regular season, take your pick!!
			ws = wb[wb.sheetnames[0]] #first sheet
		else: #playoffs (texted 'b'), second sheet
			ws = wb[wb.sheetnames[1]]
		player_and_stat = msg.split() #split 
		if len(player_and_stat) == 3: # 1 player
			full_name = player_and_stat[0] + " " + player_and_stat[1]
			stat = player_and_stat[2]
			for row in range(1, ws.max_row+1): #need +1 to get last row!
				for col in "A": #A gets players for texted season
					cell_name="{}{}".format(col, row)
					list_of_players.append(ws[cell_name].value.lower())
				for col in stat_dict[stat]: #pts
					cell_name="{}{}".format(col, row)
					list_of_stats.append(ws[cell_name].value)
			player_stat_map = dict(zip(list_of_players, list_of_stats))
			if full_name in player_stat_map.keys():
				ret = MessagingResponse().message(full_name + " averaged " + str(stat) + " of: " +str(player_stat_map[full_name]))
			else:
				ret = MessagingResponse().message(typomsg)
		elif len(player_and_stat) == 5: #2 players
			player1 = player_and_stat[0] + " " + player_and_stat[1] 
			player2 = player_and_stat[2] + " " + player_and_stat[3]
			stat = player_and_stat[4]
			for row in range(1, ws.max_row+1): #need +1 to get last row!
				for col in "A": #A gets players for texted season
					cell_name="{}{}".format(col, row)
					list_of_players.append(ws[cell_name].value.lower())
				for col in stat_dict[stat]: # gets column of whatever statistic
					cell_name="{}{}".format(col, row)
					list_of_stats.append(ws[cell_name].value)
			player_stat_map = dict(zip(list_of_players, list_of_stats))
			if player1 in player_stat_map.keys() and player2 in player_stat_map.keys():
				if player_stat_map[player1] > player_stat_map[player2]:
					ret = MessagingResponse().message(player1 + " 's total regular season " +  stat + " were " + str(player_stat_map[player1]) + ", higher than " + player2 + "\'s " + str(player_stat_map[player2]))
				else:
					ret = MessagingResponse().message(player2 + " 's total regular season " + stat + " " + str(player_stat_map[player2]) + ", higher than " + player1 + "\'s " + str(player_stat_map[player1]))
			else: #check
				ret = MessagingResponse().message("check both players' names (first and last!)")
		else: #idk how many players
			ret = MessagingResponse().message(typomsg)
	return str(ret)

if __name__ == "__main__":
	app.run(debug=True)